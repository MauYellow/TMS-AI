from docx import Document
from openai import OpenAI
import httpx
import numpy as np
from sklearn.metrics.pairwise import cosine_similarity
from flask import Flask, render_template, request, jsonify, url_for, send_from_directory
from openpyxl import load_workbook
from datetime import datetime
from dotenv import load_dotenv
import os

load_dotenv()




app = Flask(__name__)

@app.route('/')
def home():
    return render_template('index.html')

@app.route('/ask', methods=['POST'])
def ask():
    domanda = request.json.get("question")
    category = request.json.get("category")
    nome, file_path = MANUALI[category]
 
    ai_answer = chiedi_ai(domanda, f"{file_path}", nome)


    return jsonify({"answer": ai_answer})

@app.route("/download-report")
def download_report():
    return send_from_directory("Report", "Report.xlsx", as_attachment=True)

# Disabilita verifica SSL solo per test locale
http_client = httpx.Client(verify=False)

# Inizializza client OpenAI
client = OpenAI(
    api_key=os.getenv("OPENAI_APIKEY"),
    http_client=http_client
)

# Percorsi ai documenti
MANUALI = {
    "1": ("TMS", "manuale_tms.docx"),
    "2": ("Driver and Vehicles", "driver and vehicles.docx"),
    "3": ("Livechat", "prova_testo tagliato_solo prima parte.docx"),
    "4": ("POI Management", "POI Management.docx"),
    "5": ("Check In/Out and Shifts", "checkin and shift.docx"),
    "6": ("TXB Orders and Timeline", "TXB Orders and Timeline.docx"),
    "7": ("Livechat & Workgroup", "livechat and workgroup.docx")
}


# Interrogazione dell'AI con testo filtrato
def chiedi_ai(domanda, file_path, topic):
    doc = Document(file_path)
    testo = "\n".join(p.text.strip() for p in doc.paragraphs if p.text.strip())
    #print(f"Testo: {testo}")


    prompt = (
        "Rispondi basandoti sul testo del funzionamento del software TMS.\n"
        "Scrivi come un manuale, con punti numerati e uno ogni riga. Niente spiegazioni personali.\n"
        "Se esiste un punto contraddittorio, spiega entrambi.\n"
        "Se non trovi nulla, scrivi: It seems i don't have this information, try to switch category or ask another question.\n\n"
        "Se non trovi nulla, spiegami il perché non trovi nulla o se hai trovato cose simili ma hai scelto di non rispondere"
        f"Testo:\n{testo}\n\nDomanda: {domanda}\nRisposta:"
    )

    # Chiamata API
    response = client.chat.completions.create(
        model="gpt-4o",
        messages=[{"role": "user", "content": prompt}],
        temperature=0.0,
        max_tokens=600
    )

    print("\n🧠 Risposta:\n", response.choices[0].message.content)
    print(response.choices[0].message.content)
    if "it seems i" in response.choices[0].message.content.lower():
        print("Loggin the question in the excel file..")
        esito = False
        salva_excel(domanda, topic, esito, "No answer provided")
    else:
        print("Logging the answers in the excel file..")
        esito = True
        salva_excel(domanda, topic, esito, response.choices[0].message.content)
    return response.choices[0].message.content

def salva_excel(domanda, topic, esito, risposta):
    print(f"Scrivendo sul file Excel: {domanda}, {topic}")
    path_file = os.path.join("Report", "Report.xlsx")

    if not os.path.exists(path_file):
        print("❌ Il file Report.xlsx non esiste.")
        return

    wb = load_workbook(path_file)
    print("Fogli disponibili nel file:", wb.sheetnames)

    if esito == False:
      ws = wb["Unsolved Questions"]
    else:
      ws = wb["Solved Answers"]

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    try:
        ws.append([timestamp, domanda, topic, "New", risposta])
        wb.save(path_file)
        print("✅ Salvato in Report.xlsx")

    except Exception as e:
        print(f"Error: {e}")
        print("❌ Domanda non salvata in Report.xlsx")

    

#salva_domanda_excel()

# Interfaccia testuale
if __name__ == "__main__":
    app.run(debug=True)
    while True:
        print("\nQual è l'argomento?")
        print("1) TMS (Non ancora pronto)")
        print("2) Driver e Client App")
        print("3) Livechat (Non ancora pronto)")
        print("4) POI Management")
        scelta = input("Digita 1 o 2 (oppure 'q' per uscire): ").strip()

        if scelta.lower() == "q":
            break
        if scelta not in MANUALI:
            print("❌ Scelta non valida.")
            continue

        nome, file_path = MANUALI[scelta]
        print(f"\nHai scelto: {nome}")
        domanda = input("Qual è la domanda? ").strip()

        print("\n📖 Cerco la risposta nel manuale...")
        chiedi_ai(domanda, file_path, nome)
